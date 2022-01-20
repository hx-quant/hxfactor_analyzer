# -*- coding: utf-8 -*-
"""
Created on Fri Jul  2 08:36:17 2021

@author: Administrator

EXCEL类数据库，用于存储行情数据：
1.收盘价
2.总市值
3.市盈率
4.换手率
"""

import numpy as np
import pandas as pd
import openpyxl as op

FILE = 'E:\workspace_py\Database\Data_EXCEL\\'

class EDB():
    def __init__(self):
        return None
    
    def get_index(self, file):
        """
        读取行名
        """
        wb = op.load_workbook(file)
        sh = wb['Sheet1']
        cells = sh['A']
        index = list()
        for cell in cells:
            index.append(cell.value)
        if None in index:
            index.remove(None)
        wb.close()
        return index    
 
    def get_columns(self, file):
        """
        读取行名
        """
        wb = op.load_workbook(file)
        sh = wb['Sheet1']
        cells = sh['1']
        columns = list()
        for cell in cells:
            columns.append(cell.value)
        if None in columns:
            columns.remove(None)
        wb.close()
        return columns   
    
    def update(self, file, data, sheet='Sheet1', beg_idx=0, beg_col=0):
        if isinstance(data,pd.DataFrame):
            self.store_df(file, data, sheet=sheet, beg_idx=beg_idx, beg_col=beg_col)
        elif isinstance(data,list):
            self.store_list(file, data, sheet=sheet)            
        return None
    
    def store_df(self, file, data, sheet='Sheet1', beg_idx=0, beg_col=0):
        wb = op.load_workbook(file)
        if sheet not in wb.sheetnames:
            sh = wb.create_sheet(sheet)
        sh = wb[sheet]
        for i in range(0,len(data.index)):
            sh.cell(beg_idx+i+2,1,data.index[i])
        for j in range(0,len(data.columns)):
            sh.cell(1,beg_col+j+2,data.columns[j])
        for i in range(0,len(data.index)):
            for j in range(0,len(data.columns)):   
                sh.cell(beg_idx+i+2,beg_col+j+2,data.iloc[i,j])
        wb.save(file)
        return None
    
    def store_list(self, file, data, sheet='Sheet1'):
        wb = op.load_workbook(file)
        if sheet not in wb.sheetnames:
            sh = wb.create_sheet(sheet)
        sh = wb[sheet]
        for i in range(0,len(data)):
            sh.cell(i+1,1,data[i])            
        wb.save(file)
        return None
    
    def get_stocks(self):
        file = FILE + 'Information\securities_information.xlsx'
        stock_infos = pd.read_excel(file,index_col=0)
        stocks = list(stock_infos['证券代码'].values)
        return stocks
    
    def get_trade_days(self):
        file = FILE + 'Market_Data/trade_days.xlsx'
        trade_days = pd.read_excel(file)
        trade_days = list(trade_days.iloc[:,0].values)
        trade_days = [pd.to_datetime(x) for x in trade_days]
        return trade_days
    
    def get_report_dates(self):
        file = FILE + 'Information\\financial_disclosure_dates.xlsx'
        report_dates = pd.read_excel(file,header=0,index_col=0)
        report_dates = report_dates.T.iloc[1:,:]
        from Tools.myTools import change_repDates_names
        report_dates.index = change_repDates_names(report_dates.index)
        return report_dates
    
    def get_industries(self, industry='sw_l1'):
        trade_days = self.get_trade_days()
        stocks = self.get_stocks()
        industries = pd.DataFrame(index=trade_days,columns=stocks)
        for i in range(2010,2021):
            dates = list()
            for d in trade_days:
                if str(i) in d.strftime("%Y-%m-%d"):
                    dates.append(d)
            file = FILE + 'Information/industries_'+str(i)+'.xlsx'
            a = pd.read_excel(file,index_col=0)[industry]
            industries.loc[dates,a.index] = pd.DataFrame(a).T.values
        return industries
    
    def get_market_data(self, target):
        file = FILE + 'Market_Data\\'+ target + '.xlsx'
        market_data = pd.read_excel(file,header=0,index_col=0)
        return market_data

    def get_finance_data(self, target):
        clas_name = 'finance_data\\'
        file = FILE + clas_name + target + '.xlsx'
        result = pd.read_excel(file,header=0,index_col=0)
        return result